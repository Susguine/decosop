"""
Import WebDocuments from OfficeDocuments.

Reads OfficeDocuments from the database, finds the physical files in the uploads
directory, converts .docx/.doc/.xlsx to HTML (same as import_sops.py), and inserts
them as WebDocuments mapped to the corresponding WebDocCategory.

PDFs, images, and other non-convertible files are skipped.
"""

import os
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import mammoth
import openpyxl

DB_PATH = r"C:\Users\JenniferD\source\repos\DecoSOP\decosop.db"
UPLOADS_DIR = r"C:\Users\JenniferD\source\repos\DecoSOP\uploads"


def convert_docx_to_html(filepath: str) -> str | None:
    """Convert a .docx file to HTML using mammoth."""
    try:
        with open(filepath, "rb") as f:
            result = mammoth.convert_to_html(f)
            html = result.value
            if html and html.strip():
                return html
    except Exception as e:
        print(f"  WARNING: Failed to convert {filepath}: {e}")
    return None


def convert_doc_to_html(filepath: str) -> str | None:
    """Extract text from old .doc binary format (best effort)."""
    try:
        with open(filepath, "rb") as f:
            data = f.read()

        text = ""
        try:
            raw = data.decode("latin-1")
            chunks = re.findall(r"[\x20-\x7e\n\r\t]{20,}", raw)
            if chunks:
                text = max(chunks, key=len)
        except Exception:
            pass

        if not text or len(text.strip()) < 50:
            return None

        lines = text.strip().split("\n")
        html_parts = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            line = line.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            html_parts.append(f"<p>{line}</p>")

        return "\n".join(html_parts) if html_parts else None
    except Exception as e:
        print(f"  WARNING: Failed to extract text from {filepath}: {e}")
    return None


def convert_xlsx_to_html(filepath: str) -> str | None:
    """Convert an .xlsx file to an HTML table."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        all_html = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            rows = [r for r in rows if any(cell is not None and str(cell).strip() for cell in r)]
            if not rows:
                continue

            if len(wb.sheetnames) > 1:
                all_html.append(f"<h2>{sheet_name}</h2>")

            all_html.append("<table><thead><tr>")
            first_row = rows[0]
            for cell in first_row:
                val = str(cell).strip() if cell is not None else ""
                val = val.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                all_html.append(f"<th>{val}</th>")
            all_html.append("</tr></thead><tbody>")

            for row in rows[1:]:
                all_html.append("<tr>")
                for cell in row:
                    val = str(cell).strip() if cell is not None else ""
                    val = val.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    all_html.append(f"<td>{val}</td>")
                all_html.append("</tr>")
            all_html.append("</tbody></table>")

        wb.close()
        html = "\n".join(all_html)
        return html if html.strip() else None
    except Exception as e:
        print(f"  WARNING: Failed to convert {filepath}: {e}")
    return None


def build_category_map(cursor):
    """Build a mapping from DocumentCategory ID to WebDocCategory ID using full paths."""
    def build_paths(table):
        cursor.execute(f"SELECT Id, Name, ParentId FROM {table}")
        cats = {row[0]: (row[1], row[2]) for row in cursor.fetchall()}
        paths = {}
        for cid in cats:
            parts = []
            current = cid
            while current is not None:
                name, parent = cats[current]
                parts.insert(0, name)
                current = parent
            paths[cid] = "/".join(parts)
        return paths

    doc_paths = build_paths("DocumentCategories")
    web_paths = build_paths("WebDocCategories")

    path_to_web = {v: k for k, v in web_paths.items()}

    mapping = {}
    for doc_id, path in doc_paths.items():
        if path in path_to_web:
            mapping[doc_id] = path_to_web[path]
    return mapping


def main():
    if not os.path.isfile(DB_PATH):
        print(f"ERROR: Database not found: {DB_PATH}")
        print("Edit DB_PATH at the top of this script to point to your decosop.db file.")
        sys.exit(1)

    if not os.path.isdir(UPLOADS_DIR):
        print(f"ERROR: Uploads directory not found: {UPLOADS_DIR}")
        print("Edit UPLOADS_DIR at the top of this script to point to your uploads folder.")
        sys.exit(1)

    print(f"Database: {DB_PATH}")
    print(f"Uploads:  {UPLOADS_DIR}")
    print()

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Build category mapping
    cat_map = build_category_map(cursor)
    print(f"Category mapping: {len(cat_map)} categories mapped")

    # Clear existing WebDocuments
    cursor.execute("SELECT COUNT(*) FROM WebDocuments")
    existing = cursor.fetchone()[0]
    if existing > 0:
        cursor.execute("DELETE FROM WebDocuments")
        print(f"Cleared {existing} existing WebDocuments")
    print()

    # Read all OfficeDocuments
    cursor.execute("""
        SELECT Id, Title, FileName, StoredFileName, CategoryId, SortOrder
        FROM OfficeDocuments
        ORDER BY CategoryId, SortOrder
    """)
    docs = cursor.fetchall()
    print(f"Found {len(docs)} OfficeDocuments to process")

    # Count by extension
    from collections import Counter
    exts = Counter()
    for _, _, fn, _, _, _ in docs:
        exts[os.path.splitext(fn)[1].lower()] += 1
    for ext, count in exts.most_common():
        print(f"  {ext}: {count}")
    print()

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    imported = 0
    skipped = 0
    failed = 0
    unmapped_cats = 0

    # Track titles per category for uniqueness
    used_titles = {}

    for idx, (doc_id, title, filename, stored_filename, cat_id, sort_order) in enumerate(docs):
        if (idx + 1) % 200 == 0:
            print(f"  Processing {idx + 1}/{len(docs)}...")

        # Map category
        web_cat_id = cat_map.get(cat_id)
        if web_cat_id is None:
            unmapped_cats += 1
            continue

        # Check file extension
        ext = os.path.splitext(filename)[1].lower()
        if ext not in (".docx", ".doc", ".xlsx"):
            skipped += 1
            continue

        # Find the file
        filepath = os.path.join(UPLOADS_DIR, stored_filename)
        if not os.path.exists(filepath):
            failed += 1
            continue

        # Convert to HTML
        html = None
        if ext == ".docx":
            html = convert_docx_to_html(filepath)
        elif ext == ".doc":
            html = convert_doc_to_html(filepath)
        elif ext == ".xlsx":
            html = convert_xlsx_to_html(filepath)

        if html is None or len(html.strip()) < 10:
            skipped += 1
            continue

        # Wrap with title heading
        title_escaped = title.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        full_html = f"<h1>{title_escaped}</h1>\n{html}"

        # Ensure unique title within category
        if web_cat_id not in used_titles:
            used_titles[web_cat_id] = set()
        unique_title = title
        counter = 1
        while unique_title in used_titles[web_cat_id]:
            counter += 1
            unique_title = f"{title} ({counter})"
        used_titles[web_cat_id].add(unique_title)

        # Insert
        try:
            cursor.execute(
                "INSERT INTO WebDocuments (Title, HtmlContent, CategoryId, SortOrder, IsFavorited, CreatedAt, UpdatedAt) VALUES (?, ?, ?, ?, 0, ?, ?)",
                (unique_title, full_html, web_cat_id, sort_order, now, now),
            )
            imported += 1
        except sqlite3.IntegrityError as e:
            print(f"  DUPLICATE: {unique_title} in category {web_cat_id}: {e}")
            failed += 1

    conn.commit()

    print()
    print("=" * 60)
    print("Import complete!")
    print(f"  WebDocuments imported:   {imported}")
    print(f"  Skipped (non-convertible/empty): {skipped}")
    print(f"  Failed (missing file):   {failed}")
    print(f"  Unmapped categories:     {unmapped_cats}")

    cursor.execute("SELECT COUNT(*) FROM WebDocuments")
    total = cursor.fetchone()[0]
    print(f"  Total WebDocuments now:  {total}")

    conn.close()


if __name__ == "__main__":
    main()
