"""
Import SOPs from the protocols directory into the DecoSOP SQLite database.

Converts .docx files to HTML using mammoth, .xlsx to HTML tables using openpyxl,
and .doc (old binary) to plain-text HTML via basic extraction.
PDFs are skipped (no reliable text extraction without extra tools).

Creates a nested category hierarchy mirroring the directory structure on disk.
Filters out: archive, "to go thru", and "old" directories (but not "gold", "folder", etc.)
"""

import os
import re
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

import mammoth
import openpyxl

# --- Configuration ---
PROTOCOLS_DIR = r"C:\Users\JenniferD\OneDrive - ZO Organized (1)\!   Admin@deco FILES\z PROTOCOLS"
DB_PATH = r"C:\Users\JenniferD\source\repos\DecoSOP\decosop.db"

# Directories to skip (matched as standalone directory names, case-insensitive)
SKIP_PATTERNS = [
    re.compile(r"^zz\s*archive$", re.IGNORECASE),
    re.compile(r"^z\s*archive$", re.IGNORECASE),
    re.compile(r"^x\s*old$", re.IGNORECASE),
    re.compile(r"^z\s*old\b", re.IGNORECASE),
    re.compile(r"^poss\s*older\b", re.IGNORECASE),
    re.compile(r"^!+\s*.*to\s+go\s+thru", re.IGNORECASE),
    re.compile(r"to\s+go\s+thru", re.IGNORECASE),
    re.compile(r"^archive$", re.IGNORECASE),
]

# Map top-level directory names to cleaner display names
CATEGORY_MAP = {
    "@ REPORTING Protocols": "Reporting",
    "^   DENTAL  EVAL & PROTOCOLS": "Dental Evaluations & Protocols",
    "^   DENTAL  PROCEDURES": "Dental Procedures",
    "^  Hope Dental Clinic": "Hope Dental Clinic",
    "^ ORAL HEALTH Education": "Oral Health Education",
    "0 GENERAL BIZ Protocols": "General Business",
    "0 Goods & Services": "Goods & Services",
    "ACCOUNTS RECEIVABLES": "Accounts Receivable",
    "FINANCE PROTOCOLs": "Finance",
    "INSURANCE Protocols": "Insurance",
    "MARKETING Protocols": "Marketing",
    "PATIENT RELATIONS Mgmt": "Patient Relations",
    "PROPERTY Protocols": "Property",
    "REGULATION & Compliance Protocols": "Regulation & Compliance",
    "Risk MANAGEMENT": "Risk Management",
    "ROUTING Protocols": "Routing",
    "SAFETY Protocols": "Safety",
    "SCHEDULE Protocols": "Scheduling",
    "SECURITY Protocols": "Security",
    "SUPPLIES & Biz Services Protocols": "Supplies & Business Services",
    "SUPPLIES & SERVICES Protocols": "Supplies & Services",
    "Training Protocols": "Training",
    "z~~ Facial Cosmetic SERVICES PROTOCOLS": "Facial Cosmetic Services",
}

# Files at the root level (not in any subfolder) go to this category
ROOT_CATEGORY = "General"


def should_skip_dir(dirname: str) -> bool:
    """Check if a directory name matches any skip pattern."""
    for pattern in SKIP_PATTERNS:
        if pattern.search(dirname):
            return True
    return False


def clean_dirname(dirname: str, is_top_level: bool) -> str:
    """Clean a directory name into a readable category name."""
    if is_top_level:
        # Use the mapping for top-level dirs
        mapped = CATEGORY_MAP.get(dirname)
        if mapped:
            return mapped

    # General cleanup for unmapped dirs
    name = dirname
    # Remove leading special chars used for sorting (^, ~, !, @, 0-9)
    name = re.sub(r"^[~!^@]+\s*", "", name)
    name = re.sub(r"^\d+\s+", "", name)
    # Collapse multiple spaces
    name = re.sub(r"\s+", " ", name).strip()
    return name if name else dirname


def clean_title(filename: str) -> str:
    """Clean up a filename into a readable title."""
    name = Path(filename).stem
    # Remove leading special chars used for sorting
    name = re.sub(r"^[~!^]+\s*", "", name)
    # Replace underscores with spaces
    name = name.replace("_", " ")
    # Collapse multiple spaces
    name = re.sub(r"\s+", " ", name).strip()
    return name if name else filename


def convert_docx_to_html(filepath: str) -> str | None:
    """Convert a .docx file to HTML using mammoth."""
    try:
        with open(filepath, "rb") as f:
            result = mammoth.convert_to_html(f)
            html = result.value
            if html and html.strip():
                return html
    except Exception as e:
        print(f"  WARNING: Failed to convert {filepath}: {e}")
    return None


def convert_doc_to_html(filepath: str) -> str | None:
    """Extract text from old .doc binary format (best effort)."""
    try:
        with open(filepath, "rb") as f:
            data = f.read()

        text = ""
        try:
            raw = data.decode("latin-1")
            chunks = re.findall(r"[\x20-\x7e\n\r\t]{20,}", raw)
            if chunks:
                text = max(chunks, key=len)
        except Exception:
            pass

        if not text or len(text.strip()) < 50:
            return None

        lines = text.strip().split("\n")
        html_parts = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            line = line.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            html_parts.append(f"<p>{line}</p>")

        return "\n".join(html_parts) if html_parts else None
    except Exception as e:
        print(f"  WARNING: Failed to extract text from {filepath}: {e}")
    return None


def convert_xlsx_to_html(filepath: str) -> str | None:
    """Convert an .xlsx file to an HTML table."""
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        all_html = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            rows = [r for r in rows if any(cell is not None and str(cell).strip() for cell in r)]
            if not rows:
                continue

            if len(wb.sheetnames) > 1:
                all_html.append(f"<h2>{sheet_name}</h2>")

            all_html.append('<table><thead><tr>')
            first_row = rows[0]
            for cell in first_row:
                val = str(cell).strip() if cell is not None else ""
                val = val.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                all_html.append(f"<th>{val}</th>")
            all_html.append("</tr></thead><tbody>")

            for row in rows[1:]:
                all_html.append("<tr>")
                for cell in row:
                    val = str(cell).strip() if cell is not None else ""
                    val = val.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    all_html.append(f"<td>{val}</td>")
                all_html.append("</tr>")
            all_html.append("</tbody></table>")

        wb.close()
        html = "\n".join(all_html)
        return html if html.strip() else None
    except Exception as e:
        print(f"  WARNING: Failed to convert {filepath}: {e}")
    return None


def walk_protocols(base_dir: str):
    """Walk the protocols directory, yielding (rel_path, full_path, extension) tuples."""
    base = Path(base_dir)
    for root, dirs, files in os.walk(base):
        # Filter out directories to skip (modifying dirs in-place prunes the walk)
        dirs[:] = [d for d in dirs if not should_skip_dir(d)]

        for filename in sorted(files):
            ext = Path(filename).suffix.lower()
            if ext not in (".docx", ".doc", ".xlsx"):
                continue
            if filename.startswith("~$"):
                continue

            full_path = os.path.join(root, filename)
            rel_path = os.path.relpath(full_path, base)
            yield rel_path, full_path, ext


def create_tables(cursor):
    """Create the Categories and Documents tables matching the EF Core schema."""
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Categories (
            Id INTEGER PRIMARY KEY AUTOINCREMENT,
            Name TEXT NOT NULL,
            SortOrder INTEGER NOT NULL,
            ParentId INTEGER,
            FOREIGN KEY (ParentId) REFERENCES Categories(Id) ON DELETE RESTRICT
        )
    """)
    cursor.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS IX_Categories_ParentId_Name ON Categories(ParentId, Name)"
    )

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS Documents (
            Id INTEGER PRIMARY KEY AUTOINCREMENT,
            Title TEXT NOT NULL,
            HtmlContent TEXT NOT NULL,
            CategoryId INTEGER NOT NULL,
            SortOrder INTEGER NOT NULL,
            CreatedAt TEXT NOT NULL,
            UpdatedAt TEXT NOT NULL,
            FOREIGN KEY (CategoryId) REFERENCES Categories(Id) ON DELETE CASCADE
        )
    """)
    cursor.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS IX_Documents_CategoryId_Title ON Documents(CategoryId, Title)"
    )


def ensure_category(cursor, name: str, parent_id: int | None, sort_counters: dict) -> int:
    """Get or create a category, returning its Id."""
    # Check if it already exists
    if parent_id is None:
        cursor.execute(
            "SELECT Id FROM Categories WHERE Name = ? AND ParentId IS NULL", (name,)
        )
    else:
        cursor.execute(
            "SELECT Id FROM Categories WHERE Name = ? AND ParentId = ?", (name, parent_id)
        )
    row = cursor.fetchone()
    if row:
        return row[0]

    # Determine sort order among siblings
    sort_key = (parent_id,)  # group by parent
    sort_order = sort_counters.get(sort_key, 0)
    sort_counters[sort_key] = sort_order + 1

    cursor.execute(
        "INSERT INTO Categories (Name, SortOrder, ParentId) VALUES (?, ?, ?)",
        (name, sort_order, parent_id),
    )
    return cursor.lastrowid


def get_category_for_path(cursor, rel_path: str, sort_counters: dict) -> int:
    """
    Create/get the full category hierarchy for a file's relative path.
    Each directory component becomes a nested category.
    Returns the deepest category's Id.
    """
    parts = Path(rel_path).parts

    if len(parts) <= 1:
        # File is at root of PROTOCOLS_DIR — goes to "General" category
        return ensure_category(cursor, ROOT_CATEGORY, None, sort_counters)

    # Directory parts (everything except the filename)
    dir_parts = parts[:-1]

    parent_id = None
    for i, dirname in enumerate(dir_parts):
        is_top_level = (i == 0)
        clean_name = clean_dirname(dirname, is_top_level)
        parent_id = ensure_category(cursor, clean_name, parent_id, sort_counters)

    return parent_id


def main():
    print(f"Scanning: {PROTOCOLS_DIR}")
    print(f"Database: {DB_PATH}")
    print()

    # Collect all files first
    files = list(walk_protocols(PROTOCOLS_DIR))
    print(f"Found {len(files)} files to process")

    by_ext = {}
    for _, _, ext in files:
        by_ext[ext] = by_ext.get(ext, 0) + 1
    for ext, count in sorted(by_ext.items()):
        print(f"  {ext}: {count}")
    print()

    # Delete old database and create fresh
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print("Deleted old database")

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    create_tables(cursor)
    conn.commit()

    # Track sort orders: key is (parent_id,) for categories, cat_id for documents
    cat_sort_counters = {}
    doc_sort_counters = {}
    # Track titles per category to avoid duplicates
    used_titles = {}

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    doc_count = 0
    skip_count = 0
    fail_count = 0

    for idx, (rel_path, full_path, ext) in enumerate(files):
        if (idx + 1) % 100 == 0:
            print(f"  Processing {idx + 1}/{len(files)}...")

        # Get or create nested category hierarchy
        cat_id = get_category_for_path(cursor, rel_path, cat_sort_counters)

        # Build title from filename only (hierarchy handles organization)
        title = clean_title(os.path.basename(rel_path))

        # Ensure unique title within category
        if cat_id not in used_titles:
            used_titles[cat_id] = set()
        base_title = title
        counter = 1
        while title in used_titles[cat_id]:
            counter += 1
            title = f"{base_title} ({counter})"
        used_titles[cat_id].add(title)

        # Convert to HTML
        html = None
        if ext == ".docx":
            html = convert_docx_to_html(full_path)
        elif ext == ".doc":
            html = convert_doc_to_html(full_path)
        elif ext == ".xlsx":
            html = convert_xlsx_to_html(full_path)

        if html is None or len(html.strip()) < 10:
            skip_count += 1
            continue

        # Wrap with title heading
        title_escaped = title.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        full_html = f"<h1>{title_escaped}</h1>\n{html}"

        # Insert document
        sort_order = doc_sort_counters.get(cat_id, 0)
        doc_sort_counters[cat_id] = sort_order + 1
        try:
            cursor.execute(
                "INSERT INTO Documents (Title, HtmlContent, CategoryId, SortOrder, CreatedAt, UpdatedAt) VALUES (?, ?, ?, ?, ?, ?)",
                (title, full_html, cat_id, sort_order, now, now),
            )
            doc_count += 1
        except sqlite3.IntegrityError as e:
            print(f"  DUPLICATE: {title} in category {cat_id}: {e}")
            fail_count += 1

    conn.commit()

    # Print summary
    print()
    print("=" * 60)
    print(f"Import complete!")
    print(f"  Documents imported: {doc_count}")
    print(f"  Skipped (empty/failed): {skip_count}")
    print(f"  Duplicates/errors: {fail_count}")

    # Count categories
    cursor.execute("SELECT COUNT(*) FROM Categories")
    total_cats = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM Categories WHERE ParentId IS NULL")
    root_cats = cursor.fetchone()[0]
    print(f"  Categories: {total_cats} total ({root_cats} top-level)")
    print()

    # Print tree of categories with doc counts
    print("Category tree:")
    cursor.execute("""
        SELECT c.Id, c.Name, c.ParentId,
               (SELECT COUNT(*) FROM Documents d WHERE d.CategoryId = c.Id) as doc_count
        FROM Categories c
        ORDER BY c.ParentId IS NOT NULL, c.ParentId, c.SortOrder
    """)
    all_cats = cursor.fetchall()
    cat_lookup = {row[0]: row for row in all_cats}
    children_map = {}
    for cat_id, name, parent_id, doc_count_val in all_cats:
        children_map.setdefault(parent_id, []).append((cat_id, name, doc_count_val))

    def print_tree(parent_id, depth=0):
        for cat_id, name, doc_count_val in children_map.get(parent_id, []):
            indent = "  " * (depth + 1)
            suffix = f" ({doc_count_val} docs)" if doc_count_val > 0 else ""
            print(f"{indent}{name}{suffix}")
            print_tree(cat_id, depth + 1)

    print_tree(None)

    conn.close()


if __name__ == "__main__":
    main()
